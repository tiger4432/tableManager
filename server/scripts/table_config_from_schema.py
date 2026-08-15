# -*- coding: utf-8 -*-
"""Schema sheet (Table · Column · Type) -> `table_config.json` draft.

WHAT THIS DOES AND, MORE IMPORTANTLY, WHAT IT REFUSES TO DO
-----------------------------------------------------------
A schema sheet carries three facts per row: which table, which column, what the source
system calls its type. `table_config.json` needs more than that -- which column is the
business key, which columns compose it, what the table is FOR. None of those are in the
sheet, and a generator that guesses them writes a config that looks complete and is
quietly wrong. So this script fills what the sheet supports, leaves the rest EXPLICITLY
empty, and prints one report naming every hole and every guess.

Two kinds of output, and they are different things:

  * `column_types` -- filled for every column. `string` unless something says otherwise.
    A value that came from the sheet is a FACT; a value derived from the column NAME is a
    GUESS and appears in the report under 「이름으로 추정」.
  * `business_key` / `composite_key_source` / `__comment` -- left null/empty. These are
    decisions, and the report lists them as such. Candidates are SUGGESTED, never applied.

The type vocabulary is what the existing config actually uses (measured, not assumed):
`string` (132) · `number` (60) · `datetime` (3).

USAGE
-----
    python table_config_from_schema.py <sheet.xlsx|.csv|.tsv> [-o out.json] [--merge existing.json]

`--merge` keeps declarations already made for tables the sheet does not mention, and
never overwrites a non-empty `business_key` / `composite_key_source` / `__comment` that a
human already filled in -- so re-running after an edit does not undo the edit.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys

TYPE_STRING = "string"
TYPE_NUMBER = "number"
TYPE_DATETIME = "datetime"

#: Header spellings accepted for the three columns the sheet must carry. Matching is
#: case-insensitive and ignores spaces/underscores, because a sheet written by hand says
#: "Table Name" where the next one says "table_name".
HEADER_TABLE = {"table", "tablename", "table명", "테이블"}
HEADER_COLUMN = {"column", "columnname", "field", "컬럼", "필드"}
HEADER_TYPE = {"type", "datatype", "coltype", "타입", "자료형"}

#: Source-system type words that are unambiguous. Anything not listed falls through to the
#: name heuristic, which is reported as a guess.
NUMERIC_WORDS = ("int", "bigint", "smallint", "number", "numeric", "decimal", "double",
                 "float", "real", "long", "short", "byte")
DATETIME_WORDS = ("datetime", "timestamp", "date", "time")

#: Column-name endings that suggest a numeric quantity. Deliberately conservative: `_id`
#: is NOT here, because ids are routinely non-numeric even when they look numeric today,
#: and a lot id that becomes `A-01` next quarter would break a number column.
NUMERIC_NAME_SUFFIXES = ("_seq", "_no", "_num", "_cnt", "_count", "_qty", "_size",
                         "_len", "_length", "_idx", "_index", "_height", "_width",
                         "_x", "_y", "_z")
NUMERIC_NAME_EXACT = ("x", "y", "z", "seq", "qty", "count")

DATETIME_NAME_SUFFIXES = ("_time", "_at", "_date", "_dt", "_timestamp")
DATETIME_NAME_EXACT = ("time", "date", "timestamp", "eventtime", "event_time")


def _norm_header(text) -> str:
    return re.sub(r"[\s_]+", "", str(text or "")).strip().lower()


def _clean(text) -> str:
    return str(text if text is not None else "").strip()


def read_rows(path: str):
    """(table, column, type) triples from xlsx / csv / tsv. Header row required."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:                                        # pragma: no cover
            raise SystemExit("openpyxl이 없습니다. csv/tsv로 저장해서 다시 시도하십시오.")
        wb = load_workbook(path, read_only=True, data_only=True)
        grid = [[c for c in row] for row in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
    else:
        delim = "\t" if ext in (".tsv", ".txt") else ","
        with io.open(path, "r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            if "\t" in sample.splitlines()[0] if sample.splitlines() else False:
                delim = "\t"
            grid = [row for row in csv.reader(handle, delimiter=delim)]

    grid = [row for row in grid if row and any(_clean(c) for c in row)]
    if not grid:
        raise SystemExit("빈 시트입니다.")

    header = [_norm_header(c) for c in grid[0]]
    try:
        i_table = next(i for i, h in enumerate(header) if h in HEADER_TABLE)
        i_col = next(i for i, h in enumerate(header) if h in HEADER_COLUMN)
    except StopIteration:
        raise SystemExit(f"머리행에서 Table/Column 열을 찾지 못했습니다: {grid[0]}")
    i_type = next((i for i, h in enumerate(header) if h in HEADER_TYPE), None)

    out = []
    for row in grid[1:]:
        table = _clean(row[i_table] if i_table < len(row) else "")
        column = _clean(row[i_col] if i_col < len(row) else "")
        if not table or not column:
            continue
        raw_type = _clean(row[i_type]) if (i_type is not None and i_type < len(row)) else ""
        out.append((table, column, raw_type))
    if not out:
        raise SystemExit("데이터 행이 없습니다.")
    return out


def infer_type(column: str, raw_type: str):
    """-> (type, basis) where basis is 'sheet' (a fact) or 'name' (a guess) or 'default'."""
    low_type = raw_type.strip().lower()
    if any(w in low_type for w in DATETIME_WORDS):
        return TYPE_DATETIME, "sheet"
    if any(w in low_type for w in NUMERIC_WORDS):
        return TYPE_NUMBER, "sheet"

    name = column.strip().lower()
    if name in DATETIME_NAME_EXACT or name.endswith(DATETIME_NAME_SUFFIXES):
        return TYPE_DATETIME, "name"
    if name in NUMERIC_NAME_EXACT or name.endswith(NUMERIC_NAME_SUFFIXES):
        return TYPE_NUMBER, "name"
    return TYPE_STRING, "default"


def key_candidates(table: str, columns):
    """SUGGESTIONS for `business_key`, never applied. Ordered most-specific first."""
    low = {c.lower(): c for c in columns}
    hits = []
    for want in (f"{table.lower()}_id", "business_key_val", "row_id", "id"):
        if want in low:
            hits.append(low[want])
    hits += [c for c in columns if c.lower().endswith("_id") and c not in hits]
    return hits[:4]


def build(rows, existing=None):
    existing = existing or {}
    tables = {}
    order = []
    for table, column, raw in rows:
        if table not in tables:
            tables[table] = []
            order.append(table)
        tables[table].append((column, raw))

    config = dict(existing)
    guesses, decisions = [], []

    for table in order:
        cols = tables[table]
        prior = existing.get(table) or {}
        column_types = {}
        for column, raw in cols:
            kind, basis = infer_type(column, raw)
            column_types[column] = kind
            if basis == "name":
                guesses.append((table, column, kind, raw or "(타입 칸 비어 있음)"))

        decl = {
            # A comment is what tells the next reader what the table is FOR, and no sheet
            # carries that. Empty on purpose; the report asks for it.
            "__comment": prior.get("__comment", ""),
            # The single decision this generator must not make.
            "business_key": prior.get("business_key"),
            "column_types": column_types,
            # Sheet order is the honest default: it loses nothing and hides nothing.
            "display_columns": [c for c, _ in cols],
        }
        for optional in ("composite_key_source", "composite_key_separator", "map_key_columns"):
            if optional in prior:
                decl[optional] = prior[optional]

        if not decl["business_key"]:
            decisions.append((table, "business_key", key_candidates(table, [c for c, _ in cols])))
        if not decl["__comment"]:
            decisions.append((table, "__comment", []))

        config[table] = decl

    return config, guesses, decisions, order


def report(order, guesses, decisions, out_path):
    lines = []
    lines.append(f"테이블 {len(order)}개 초안 작성 → {out_path}")
    lines.append("")
    if guesses:
        lines.append(f"[이름으로 추정한 타입 {len(guesses)}건] — 시트가 말하지 않아 컬럼 이름으로 정했습니다. 틀렸으면 고치십시오.")
        for table, column, kind, raw in guesses:
            lines.append(f"  · {table}.{column} → {kind}   (시트 타입: {raw})")
        lines.append("")
    else:
        lines.append("[추정 없음] 모든 타입이 시트에서 나왔거나 기본값 string입니다.")
        lines.append("")

    need_key = [(t, c) for t, field, c in decisions if field == "business_key"]
    if need_key:
        lines.append(f"[사람이 정해야 함 — business_key {len(need_key)}개] 이 스크립트는 채우지 않습니다.")
        for table, cands in need_key:
            hint = ", ".join(cands) if cands else "후보 없음 — 복합 키가 필요할 수 있습니다"
            lines.append(f"  · {table}: 후보 {hint}")
        lines.append("")
    need_comment = [t for t, field, _ in decisions if field == "__comment"]
    if need_comment:
        lines.append(f"[사람이 정해야 함 — __comment {len(need_comment)}개] 이 표가 «무엇인지» 한 줄. 없으면 다음 사람이 컬럼 이름으로 추측하게 됩니다.")
        lines.append("  · " + ", ".join(need_comment))
        lines.append("")
    lines.append("한 행이 여러 컬럼으로 합쳐진 키를 쓰면 composite_key_source / composite_key_separator를 손으로 넣으십시오 (예: [\"lot\",\"event_type\",\"event_time\"] · \"|\").")
    return "\n".join(lines)


def main(argv=None):
    ap = argparse.ArgumentParser(description="스키마 시트에서 table_config 초안을 만든다")
    ap.add_argument("sheet", help="xlsx / csv / tsv (머리행에 Table·Column·Type)")
    ap.add_argument("-o", "--out", default="table_config.draft.json")
    ap.add_argument("--merge", default=None,
                    help="기존 table_config.json — 사람이 채운 값은 덮어쓰지 않는다")
    args = ap.parse_args(argv)

    rows = read_rows(args.sheet)
    existing = {}
    if args.merge and os.path.exists(args.merge):
        with io.open(args.merge, "r", encoding="utf-8") as handle:
            existing = json.load(handle) or {}

    config, guesses, decisions, order = build(rows, existing)
    with io.open(args.out, "w", encoding="utf-8") as handle:
        json.dump(config, handle, ensure_ascii=False, indent=2)
        handle.write("\n")

    text = report(order, guesses, decisions, args.out)
    try:
        print(text)
    except UnicodeEncodeError:                                     # pragma: no cover
        sys.stdout.buffer.write(text.encode("utf-8", "replace"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
